import streamlit as st
import openpyxl
import requests
import hashlib
import random
import time
from io import BytesIO

# 语言判断函数
def is_chinese(text):
    for ch in text:
        if '\u4e00' <= ch <= '\u9fff':
            return True
    return False

def is_english(text):
    for ch in text:
        if ch.isalpha():
            return True
    return False

# 读取有道API密钥
YOUDAO_APP_KEY = st.secrets["YOUDAO_APP_KEY"]
YOUDAO_APP_SECRET = st.secrets["YOUDAO_APP_SECRET"]

# 有道翻译API调用
def youdao_translate(text, from_lang='auto', to_lang='en'):
    url = 'https://openapi.youdao.com/api'
    salt = str(random.randint(1, 65536))
    curtime = str(int(time.time()))
    sign_str = YOUDAO_APP_KEY + truncate(text) + salt + curtime + YOUDAO_APP_SECRET
    sign = hashlib.sha256(sign_str.encode('utf-8')).hexdigest()
    params = {
        'q': text,
        'from': from_lang,
        'to': to_lang,
        'appKey': YOUDAO_APP_KEY,
        'salt': salt,
        'sign': sign,
        'signType': 'v3',
        'curtime': curtime,
    }
    try:
        response = requests.post(url, data=params, timeout=5)
        result = response.json()
        st.write(f"原文: {text}；返回: {result}")
        if result.get('errorCode') == '0' and 'translation' in result:
            return result['translation'][0]
        else:
            return f"[翻译失败: 错误码{result.get('errorCode')}]"
    except Exception as e:
        st.write(f"请求异常: {e}")
        return '[翻译失败]'

def truncate(text):
    if text is None:
        return ''
    size = len(text)
    return text if size <= 20 else text[:10] + str(size) + text[-10:]

st.title("Excel 批量中英文对照翻译工具（有道API版）")

uploaded_file = st.file_uploader("上传 Excel 文件", type=["xlsx"])

if uploaded_file:
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active
    st.write("正在翻译，请稍候...")

    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                text = str(cell.value).strip()
                # 优先判断中文
                if is_chinese(text):
                    to_lang = 'en'
                # 再判断英文
                elif is_english(text):
                    to_lang = 'zh-CHS'
                else:
                    # 其它内容（如数字、符号）也尝试翻译成英文
                    to_lang = 'en'
                translation = youdao_translate(text, to_lang=to_lang)
                st.write(f"单元格原文: {text}，翻译结果: {translation}")
                cell.value = f"{cell.value}\n{translation}"
                for row in ws.iter_rows():
    for cell in row:
        if cell.value:
            text = str(cell.value).strip()
            if is_chinese(text):
                to_lang = 'en'
            elif is_english(text):
                to_lang = 'zh-CHS'
            else:
                to_lang = 'en'
            translation = youdao_translate(text, to_lang=to_lang)
            st.write(f"单元格原文: {text}，翻译结果: {translation}")
            cell.value = f"{cell.value}\n{translation}"
            time.sleep(1)  # 新增：每次翻译后暂停1秒，防止被限流

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    st.success("翻译完成！")
    st.download_button(
        label="下载翻译后的Excel文件",
        data=output.getvalue(),
        file_name="output_with_translation.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
